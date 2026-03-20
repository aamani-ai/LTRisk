"""
Export an interactive Excel workbook demonstrating that
Area Under Exceedance Curve = Mean of the Data.

This is the key identity behind SCVR:
  SCVR = (area_future - area_baseline) / area_baseline
       = (mean_future - mean_baseline) / mean_baseline

Usage:
    python scripts/utils/export_exceedance_demo.py
"""

from pathlib import Path

import numpy as np
import xarray as xr
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
CACHE_DIR = Path(__file__).resolve().parents[2] / "data" / "cache" / "thredds"
OUTPUT_DIR = Path(__file__).resolve().parents[2] / "data" / "output" / "demo"
OUTPUT_FILE = OUTPUT_DIR / "exceedance_demo_tasmax.xlsx"

LAT, LON = "31.8160", "-104.0853"
VARIABLE = "tasmax"
BASELINE_YEAR = 2000
FUTURE_YEAR = 2040
SCENARIO = "ssp245"

MODELS = [
    "ACCESS-CM2",
    "ACCESS-ESM1-5",
    "BCC-CSM2-MR",
    "CMCC-ESM2",
    "CNRM-CM6-1",
    "CanESM5",
    "EC-Earth3",
    "FGOALS-g3",
    "GFDL-ESM4",
    "GISS-E2-1-G",
]

# Styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUM_FMT_2 = "0.00"
NUM_FMT_4 = "0.0000"
NUM_FMT_6 = "0.000000"
NUM_FMT_PCT = "0.0000%"


# ── Data Loading ────────────────────────────────────────────────────────────

def load_nc(model: str, scenario: str, year: int) -> np.ndarray:
    """Load a single NetCDF file, return daily values in Celsius."""
    fname = f"{model}_{scenario}_{VARIABLE}_{year}_{LAT}_{LON}.nc"
    path = CACHE_DIR / fname
    if not path.exists():
        raise FileNotFoundError(path)
    ds = xr.open_dataset(path)
    vals = ds[VARIABLE].values.flatten()
    ds.close()
    # Kelvin -> Celsius
    if np.nanmean(vals) > 100:
        vals = vals - 273.15
    return vals


def load_all_data():
    """Load baseline + future data for all models. Returns dict of arrays."""
    baseline = {}
    future = {}
    for model in MODELS:
        try:
            baseline[model] = load_nc(model, "historical", BASELINE_YEAR)
            future[model] = load_nc(model, SCENARIO, FUTURE_YEAR)
        except FileNotFoundError as e:
            print(f"  Skipping {model}: {e}")
    return baseline, future


# ── Helpers ─────────────────────────────────────────────────────────────────

def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=30):
    """Auto-fit column widths."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col[:50]:  # sample first 50 rows
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def write_label(ws, row, col, text, font=None, fill=None):
    """Write a styled label cell."""
    cell = ws.cell(row=row, column=col, value=text)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.border = THIN_BORDER
    return cell


def write_value(ws, row, col, value, fmt=NUM_FMT_2, fill=None):
    """Write a styled numeric cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill
    return cell


# ── Sheet Builders ──────────────────────────────────────────────────────────

def build_readme(wb):
    """Sheet 1: README — what this workbook demonstrates."""
    ws = wb.active
    ws.title = "README"
    ws.sheet_properties.tabColor = "2F5496"

    lines = [
        ("EXCEEDANCE CURVE DEMO", Font(bold=True, size=16, color="2F5496")),
        ("", None),
        ("KEY IDENTITY:", Font(bold=True, size=13)),
        ("  Area under the exceedance curve  =  Mean of the data", Font(size=12, color="C00000")),
        ("", None),
        ("WHY THIS MATTERS FOR SCVR:", Font(bold=True, size=13)),
        ("  SCVR = (area_future - area_baseline) / area_baseline", Font(size=12)),
        ("       = (mean_future - mean_baseline) / mean_baseline", Font(size=12, color="C00000")),
        ("", None),
        ("WHAT IS AN EXCEEDANCE CURVE?", Font(bold=True, size=13)),
        ("  1. Take all your data values (e.g. daily max temperatures)", None),
        ("  2. Sort them from LARGEST to SMALLEST", None),
        ("  3. Assign each a probability: 1st value = 0%, last = 100%", None),
        ("  4. Plot: X-axis = probability, Y-axis = value", None),
        ("  5. The AREA under this curve equals the MEAN of the data!", None),
        ("", None),
        ("WHY? (Intuition)", Font(bold=True, size=13)),
        ("  The area under the exceedance curve is the integral of the", None),
        ("  inverse CDF from 0 to 1. By a fundamental property of", None),
        ("  probability, this integral equals the expected value (mean).", None),
        ("  In discrete terms: the trapezoid approximation of this area", None),
        ("  converges to the arithmetic mean as N grows.", None),
        ("", None),
        ("SHEETS IN THIS WORKBOOK:", Font(bold=True, size=13)),
        ("  1. README (this sheet) — overview", None),
        ("  2. Raw Daily Data — actual CMIP6 temperatures (10 models)", None),
        ("  3. Exceedance Curve Data — sorted values + trapezoid areas", None),
        ("  4. Small Example — 20 values, hand-traceable proof", None),
        ("  5. SCVR Calculation — area-based vs mean-based (same!)", None),
        ("  6. Experiment Zone — change values, watch area = mean", None),
        ("", None),
        ("DATA:", Font(bold=True, size=13)),
        (f"  Variable: {VARIABLE} (daily maximum temperature, Celsius)", None),
        (f"  Site: Hayhurst Texas Solar ({LAT}, {LON})", None),
        (f"  Baseline year: {BASELINE_YEAR} (historical)", None),
        (f"  Future year: {FUTURE_YEAR} (SSP2-4.5)", None),
        (f"  Models: {len(MODELS)} CMIP6 models", None),
    ]

    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font

    ws.column_dimensions["A"].width = 80


def build_raw_data(wb, baseline, future):
    """Sheet 2: Raw Daily Data — all daily values with model labels."""
    ws = wb.create_sheet("Raw Daily Data")
    ws.sheet_properties.tabColor = "4472C4"

    # Summary stats at top
    all_base = np.concatenate(list(baseline.values()))
    all_fut = np.concatenate(list(future.values()))

    ws.cell(row=1, column=1, value="SUMMARY STATISTICS").font = Font(bold=True, size=13, color="2F5496")

    headers_summary = ["", "Count", "Mean (C)", "Min (C)", "Max (C)", "Std Dev"]
    for c, h in enumerate(headers_summary, 1):
        write_label(ws, 2, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    write_label(ws, 3, 1, f"Baseline ({BASELINE_YEAR})", fill=ACCENT_FILL)
    for c, v in enumerate([len(all_base), np.mean(all_base), np.min(all_base), np.max(all_base), np.std(all_base)], 2):
        write_value(ws, 3, c, float(v), fmt=NUM_FMT_2, fill=ACCENT_FILL)

    write_label(ws, 4, 1, f"Future ({FUTURE_YEAR}, SSP2-4.5)", fill=ACCENT_FILL)
    for c, v in enumerate([len(all_fut), np.mean(all_fut), np.min(all_fut), np.max(all_fut), np.std(all_fut)], 2):
        write_value(ws, 4, c, float(v), fmt=NUM_FMT_2, fill=ACCENT_FILL)

    # Raw data table
    data_start = 7
    headers = ["Row #", "Model", "Scenario", "Day of Year", "tasmax (C)"]
    for c, h in enumerate(headers, 1):
        write_label(ws, data_start, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    row = data_start + 1
    for model in sorted(baseline.keys()):
        vals = baseline[model]
        for day_idx, v in enumerate(vals):
            ws.cell(row=row, column=1, value=row - data_start)
            ws.cell(row=row, column=2, value=model)
            ws.cell(row=row, column=3, value=f"Baseline ({BASELINE_YEAR})")
            ws.cell(row=row, column=4, value=day_idx + 1)
            c5 = ws.cell(row=row, column=5, value=round(float(v), 2))
            c5.number_format = NUM_FMT_2
            row += 1

    for model in sorted(future.keys()):
        vals = future[model]
        for day_idx, v in enumerate(vals):
            ws.cell(row=row, column=1, value=row - data_start)
            ws.cell(row=row, column=2, value=model)
            ws.cell(row=row, column=3, value=f"Future ({FUTURE_YEAR}, SSP2-4.5)")
            ws.cell(row=row, column=4, value=day_idx + 1)
            c5 = ws.cell(row=row, column=5, value=round(float(v), 2))
            c5.number_format = NUM_FMT_2
            row += 1

    # Freeze panes and auto-filter
    ws.freeze_panes = f"A{data_start + 1}"
    ws.auto_filter.ref = f"A{data_start}:E{row - 1}"
    auto_width(ws)


def build_exceedance_curve(wb, baseline, future):
    """Sheet 3: Full exceedance curve data + chart."""
    ws = wb.create_sheet("Exceedance Curve Data")
    ws.sheet_properties.tabColor = "ED7D31"

    all_base = np.sort(np.concatenate(list(baseline.values())))[::-1]
    all_fut = np.sort(np.concatenate(list(future.values())))[::-1]

    n_b = len(all_base)
    n_f = len(all_fut)
    exc_b = np.linspace(0, 1, n_b)
    exc_f = np.linspace(0, 1, n_f)

    # Title
    ws.cell(row=1, column=1, value="EXCEEDANCE CURVE: Area Under Curve = Mean").font = Font(bold=True, size=13, color="2F5496")

    # ── Baseline table (columns A-E) ──
    ws.cell(row=3, column=1, value=f"BASELINE ({BASELINE_YEAR})").font = Font(bold=True, size=12, color="2F5496")
    b_headers = ["Rank", "Value (C)", "Exceedance Prob", "Trapezoid Width", "Trapezoid Area"]
    for c, h in enumerate(b_headers, 1):
        write_label(ws, 4, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    for i in range(n_b):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=round(float(all_base[i]), 4)).number_format = NUM_FMT_4
        ws.cell(row=r, column=3, value=round(float(exc_b[i]), 6)).number_format = NUM_FMT_6
        # Trapezoid width
        if i == 0:
            ws.cell(row=r, column=4, value=0).number_format = NUM_FMT_6
            ws.cell(row=r, column=5, value=0).number_format = NUM_FMT_6
        else:
            width = float(exc_b[i] - exc_b[i - 1])
            area = width * (float(all_base[i - 1]) + float(all_base[i])) / 2
            ws.cell(row=r, column=4, value=round(width, 8)).number_format = NUM_FMT_6
            ws.cell(row=r, column=5, value=round(area, 8)).number_format = NUM_FMT_6

    b_end = 4 + n_b  # last data row

    # Summary row for baseline
    sr = b_end + 2
    write_label(ws, sr, 4, "SUM (= Area):", font=Font(bold=True), fill=RESULT_FILL)
    sum_cell = ws.cell(row=sr, column=5)
    sum_cell.value = f"=SUM(E5:E{b_end})"
    sum_cell.number_format = NUM_FMT_4
    sum_cell.fill = RESULT_FILL
    sum_cell.font = Font(bold=True)

    write_label(ws, sr + 1, 4, "AVERAGE(values):", font=Font(bold=True), fill=HIGHLIGHT_FILL)
    avg_cell = ws.cell(row=sr + 1, column=5)
    avg_cell.value = f"=AVERAGE(B5:B{b_end})"
    avg_cell.number_format = NUM_FMT_4
    avg_cell.fill = HIGHLIGHT_FILL
    avg_cell.font = Font(bold=True)

    write_label(ws, sr + 2, 4, "Difference:", font=Font(bold=True))
    diff_cell = ws.cell(row=sr + 2, column=5)
    diff_cell.value = f"=E{sr}-E{sr+1}"
    diff_cell.number_format = NUM_FMT_6
    diff_cell.font = Font(bold=True, color="C00000")

    # ── Future table (columns G-K) ──
    ws.cell(row=3, column=7, value=f"FUTURE ({FUTURE_YEAR}, SSP2-4.5)").font = Font(bold=True, size=12, color="C00000")
    f_headers = ["Rank", "Value (C)", "Exceedance Prob", "Trapezoid Width", "Trapezoid Area"]
    for c, h in enumerate(f_headers, 7):
        write_label(ws, 4, c, h, font=HEADER_FONT, fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"))

    for i in range(n_f):
        r = 5 + i
        ws.cell(row=r, column=7, value=i + 1)
        ws.cell(row=r, column=8, value=round(float(all_fut[i]), 4)).number_format = NUM_FMT_4
        ws.cell(row=r, column=9, value=round(float(exc_f[i]), 6)).number_format = NUM_FMT_6
        if i == 0:
            ws.cell(row=r, column=10, value=0).number_format = NUM_FMT_6
            ws.cell(row=r, column=11, value=0).number_format = NUM_FMT_6
        else:
            width = float(exc_f[i] - exc_f[i - 1])
            area = width * (float(all_fut[i - 1]) + float(all_fut[i])) / 2
            ws.cell(row=r, column=10, value=round(width, 8)).number_format = NUM_FMT_6
            ws.cell(row=r, column=11, value=round(area, 8)).number_format = NUM_FMT_6

    f_end = 4 + n_f

    # Summary row for future
    fr = f_end + 2
    # Put future summary at same row as baseline summary for neatness
    sr_f = sr  # use same rows
    write_label(ws, sr_f, 10, "SUM (= Area):", font=Font(bold=True), fill=RESULT_FILL)
    sum_f = ws.cell(row=sr_f, column=11)
    sum_f.value = f"=SUM(K5:K{f_end})"
    sum_f.number_format = NUM_FMT_4
    sum_f.fill = RESULT_FILL
    sum_f.font = Font(bold=True)

    write_label(ws, sr_f + 1, 10, "AVERAGE(values):", font=Font(bold=True), fill=HIGHLIGHT_FILL)
    avg_f = ws.cell(row=sr_f + 1, column=11)
    avg_f.value = f"=AVERAGE(H5:H{f_end})"
    avg_f.number_format = NUM_FMT_4
    avg_f.fill = HIGHLIGHT_FILL
    avg_f.font = Font(bold=True)

    write_label(ws, sr_f + 2, 10, "Difference:", font=Font(bold=True))
    diff_f = ws.cell(row=sr_f + 2, column=11)
    diff_f.value = f"=K{sr_f}-K{sr_f+1}"
    diff_f.number_format = NUM_FMT_6
    diff_f.font = Font(bold=True, color="C00000")

    # ── Chart: Exceedance curves overlaid ──
    # Sample every Nth point for chart (plotting 3000+ is slow)
    chart_sample = max(1, min(n_b, n_f) // 200)

    # Write sampled data for chart in columns M-O
    ws.cell(row=4, column=13, value="Exc Prob").font = HEADER_FONT
    ws.cell(row=4, column=14, value="Baseline (C)").font = HEADER_FONT
    ws.cell(row=4, column=15, value="Future (C)").font = HEADER_FONT

    chart_rows = 0
    for i in range(0, min(n_b, n_f), chart_sample):
        r = 5 + chart_rows
        ws.cell(row=r, column=13, value=round(float(exc_b[i]), 4))
        ws.cell(row=r, column=14, value=round(float(all_base[i]), 2))
        ws.cell(row=r, column=15, value=round(float(all_fut[i]), 2))
        chart_rows += 1

    chart_end = 4 + chart_rows

    chart = LineChart()
    chart.title = "Exceedance Curves: Baseline vs Future"
    chart.x_axis.title = "Exceedance Probability"
    chart.y_axis.title = "tasmax (C)"
    chart.style = 10
    chart.width = 28
    chart.height = 16

    cats = Reference(ws, min_col=13, min_row=5, max_row=chart_end)
    base_series = Reference(ws, min_col=14, min_row=4, max_row=chart_end)
    fut_series = Reference(ws, min_col=15, min_row=4, max_row=chart_end)

    chart.add_data(base_series, titles_from_data=True)
    chart.add_data(fut_series, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.width = 25000
    chart.series[1].graphicalProperties.line.width = 25000

    ws.add_chart(chart, "M6")
    auto_width(ws, max_width=18)


def build_small_example(wb, baseline, future):
    """Sheet 4: 20 values — hand-traceable proof."""
    ws = wb.create_sheet("Small Example (20 values)")
    ws.sheet_properties.tabColor = "70AD47"

    all_base = np.sort(np.concatenate(list(baseline.values())))[::-1]
    all_fut = np.sort(np.concatenate(list(future.values())))[::-1]

    # Sample 20 evenly-spaced values
    n_sample = 20
    b_idx = np.linspace(0, len(all_base) - 1, n_sample, dtype=int)
    f_idx = np.linspace(0, len(all_fut) - 1, n_sample, dtype=int)
    b_sample = all_base[b_idx]
    f_sample = all_fut[f_idx]

    ws.cell(row=1, column=1, value="SMALL EXAMPLE: 20 values — verify by hand!").font = Font(bold=True, size=13, color="2F5496")
    ws.cell(row=2, column=1, value="The SUM of trapezoid areas equals the AVERAGE of the values. Try it!").font = Font(italic=True, size=11)

    # ── Baseline (cols A-F) ──
    ws.cell(row=4, column=1, value=f"BASELINE (20 samples from {BASELINE_YEAR})").font = Font(bold=True, size=12, color="2F5496")
    b_headers = ["Rank", "Value (C)", "Exc. Prob", "Trap. Width", "Trap. Area", "Formula"]
    for c, h in enumerate(b_headers, 1):
        write_label(ws, 5, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    exc_probs = np.linspace(0, 1, n_sample)
    for i in range(n_sample):
        r = 6 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=round(float(b_sample[i]), 2)).number_format = NUM_FMT_2
        ws.cell(row=r, column=3, value=round(float(exc_probs[i]), 4)).number_format = NUM_FMT_4

        if i == 0:
            ws.cell(row=r, column=4, value=0).number_format = NUM_FMT_4
            ws.cell(row=r, column=5, value=0).number_format = NUM_FMT_4
            ws.cell(row=r, column=6, value="(first point, no area)")
        else:
            # Use Excel formulas so they can see the logic
            ws.cell(row=r, column=4).value = f"=C{r}-C{r-1}"
            ws.cell(row=r, column=4).number_format = NUM_FMT_4
            ws.cell(row=r, column=5).value = f"=D{r}*(B{r-1}+B{r})/2"
            ws.cell(row=r, column=5).number_format = NUM_FMT_4
            ws.cell(row=r, column=6, value=f"width * (val[{i}] + val[{i+1}]) / 2")

    b_end = 5 + n_sample

    # Summary
    sr = b_end + 1
    write_label(ws, sr, 4, "SUM (Area):", font=Font(bold=True, size=12), fill=RESULT_FILL)
    ws.cell(row=sr, column=5, value=f"=SUM(E6:E{b_end})").number_format = NUM_FMT_4
    ws.cell(row=sr, column=5).fill = RESULT_FILL
    ws.cell(row=sr, column=5).font = Font(bold=True, size=12)

    write_label(ws, sr + 1, 4, "AVERAGE:", font=Font(bold=True, size=12), fill=HIGHLIGHT_FILL)
    ws.cell(row=sr + 1, column=5, value=f"=AVERAGE(B6:B{b_end})").number_format = NUM_FMT_4
    ws.cell(row=sr + 1, column=5).fill = HIGHLIGHT_FILL
    ws.cell(row=sr + 1, column=5).font = Font(bold=True, size=12)

    write_label(ws, sr + 2, 4, "Match?", font=Font(bold=True, size=12, color="C00000"))
    ws.cell(row=sr + 2, column=5, value=f'=IF(ABS(E{sr}-E{sr+1})<0.5,"YES - they match!","NO - check data")').font = Font(bold=True, size=12, color="C00000")

    # ── Future (cols H-M) ──
    ws.cell(row=4, column=8, value=f"FUTURE (20 samples from {FUTURE_YEAR})").font = Font(bold=True, size=12, color="C00000")
    f_headers = ["Rank", "Value (C)", "Exc. Prob", "Trap. Width", "Trap. Area", "Formula"]
    for c, h in enumerate(f_headers, 8):
        write_label(ws, 5, c, h, font=HEADER_FONT, fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"))

    for i in range(n_sample):
        r = 6 + i
        ws.cell(row=r, column=8, value=i + 1)
        ws.cell(row=r, column=9, value=round(float(f_sample[i]), 2)).number_format = NUM_FMT_2
        ws.cell(row=r, column=10, value=round(float(exc_probs[i]), 4)).number_format = NUM_FMT_4

        if i == 0:
            ws.cell(row=r, column=11, value=0).number_format = NUM_FMT_4
            ws.cell(row=r, column=12, value=0).number_format = NUM_FMT_4
            ws.cell(row=r, column=13, value="(first point, no area)")
        else:
            ws.cell(row=r, column=11).value = f"=J{r}-J{r-1}"
            ws.cell(row=r, column=11).number_format = NUM_FMT_4
            ws.cell(row=r, column=12).value = f"=K{r}*(I{r-1}+I{r})/2"
            ws.cell(row=r, column=12).number_format = NUM_FMT_4
            ws.cell(row=r, column=13, value=f"width * (val[{i}] + val[{i+1}]) / 2")

    f_end = 5 + n_sample

    # Future summary
    write_label(ws, sr, 11, "SUM (Area):", font=Font(bold=True, size=12), fill=RESULT_FILL)
    ws.cell(row=sr, column=12, value=f"=SUM(L6:L{f_end})").number_format = NUM_FMT_4
    ws.cell(row=sr, column=12).fill = RESULT_FILL
    ws.cell(row=sr, column=12).font = Font(bold=True, size=12)

    write_label(ws, sr + 1, 11, "AVERAGE:", font=Font(bold=True, size=12), fill=HIGHLIGHT_FILL)
    ws.cell(row=sr + 1, column=12, value=f"=AVERAGE(I6:I{f_end})").number_format = NUM_FMT_4
    ws.cell(row=sr + 1, column=12).fill = HIGHLIGHT_FILL
    ws.cell(row=sr + 1, column=12).font = Font(bold=True, size=12)

    write_label(ws, sr + 2, 11, "Match?", font=Font(bold=True, size=12, color="C00000"))
    ws.cell(row=sr + 2, column=12, value=f'=IF(ABS(L{sr}-L{sr+1})<0.5,"YES - they match!","NO - check data")').font = Font(bold=True, size=12, color="C00000")

    # ── Chart: area chart showing trapezoid slices ──
    chart = AreaChart()
    chart.title = "Exceedance Curve with Trapezoid Slices (Baseline)"
    chart.x_axis.title = "Exceedance Probability"
    chart.y_axis.title = "tasmax (C)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=3, min_row=6, max_row=b_end)
    vals = Reference(ws, min_col=2, min_row=5, max_row=b_end)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "A" + str(sr + 5))

    # Second chart for future
    chart2 = AreaChart()
    chart2.title = "Exceedance Curve with Trapezoid Slices (Future)"
    chart2.x_axis.title = "Exceedance Probability"
    chart2.y_axis.title = "tasmax (C)"
    chart2.style = 10
    chart2.width = 22
    chart2.height = 14

    cats2 = Reference(ws, min_col=10, min_row=6, max_row=f_end)
    vals2 = Reference(ws, min_col=9, min_row=5, max_row=f_end)
    chart2.add_data(vals2, titles_from_data=True)
    chart2.set_categories(cats2)

    ws.add_chart(chart2, "H" + str(sr + 5))

    auto_width(ws, max_width=18)


def build_scvr_calculation(wb, baseline, future):
    """Sheet 5: SCVR Calculation — area vs mean, per-model breakdown."""
    ws = wb.create_sheet("SCVR Calculation")
    ws.sheet_properties.tabColor = "FFC000"

    all_base = np.sort(np.concatenate(list(baseline.values())))[::-1]
    all_fut = np.sort(np.concatenate(list(future.values())))[::-1]

    mean_b = float(np.mean(all_base))
    mean_f = float(np.mean(all_fut))

    exc_b = np.linspace(0, 1, len(all_base))
    exc_f = np.linspace(0, 1, len(all_fut))
    area_b = float(np.trapezoid(all_base, exc_b))
    area_f = float(np.trapezoid(all_fut, exc_f))

    scvr_mean = (mean_f - mean_b) / mean_b
    scvr_area = (area_f - area_b) / area_b

    # Title
    ws.cell(row=1, column=1, value="SCVR CALCULATION: Two Methods, Same Answer").font = Font(bold=True, size=14, color="2F5496")

    # ── Method comparison table ──
    ws.cell(row=3, column=1, value="THE PROOF").font = Font(bold=True, size=13, color="2F5496")

    headers = ["Method", "Baseline Value", "Future Value", "SCVR = (Future - Base) / Base"]
    for c, h in enumerate(headers, 1):
        write_label(ws, 4, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    # Row 1: Via means
    write_label(ws, 5, 1, "Via MEAN", font=Font(bold=True), fill=RESULT_FILL)
    write_value(ws, 5, 2, mean_b, fmt=NUM_FMT_4, fill=RESULT_FILL)
    write_value(ws, 5, 3, mean_f, fmt=NUM_FMT_4, fill=RESULT_FILL)
    ws.cell(row=5, column=4, value=f"=(C5-B5)/B5").number_format = NUM_FMT_PCT
    ws.cell(row=5, column=4).fill = RESULT_FILL
    ws.cell(row=5, column=4).font = Font(bold=True, size=12)

    # Row 2: Via areas
    write_label(ws, 6, 1, "Via EXCEEDANCE AREA", font=Font(bold=True), fill=HIGHLIGHT_FILL)
    write_value(ws, 6, 2, area_b, fmt=NUM_FMT_4, fill=HIGHLIGHT_FILL)
    write_value(ws, 6, 3, area_f, fmt=NUM_FMT_4, fill=HIGHLIGHT_FILL)
    ws.cell(row=6, column=4, value=f"=(C6-B6)/B6").number_format = NUM_FMT_PCT
    ws.cell(row=6, column=4).fill = HIGHLIGHT_FILL
    ws.cell(row=6, column=4).font = Font(bold=True, size=12)

    # Row 3: Difference
    write_label(ws, 7, 1, "Difference", font=Font(bold=True, color="C00000"))
    write_value(ws, 7, 2, abs(mean_b - area_b), fmt=NUM_FMT_6)
    write_value(ws, 7, 3, abs(mean_f - area_f), fmt=NUM_FMT_6)
    ws.cell(row=7, column=4, value=f"=ABS(D5-D6)").number_format = NUM_FMT_6
    ws.cell(row=7, column=4).font = Font(bold=True, color="C00000")

    ws.cell(row=8, column=1, value="(Tiny difference is due to trapezoid approximation — vanishes with more points)").font = Font(italic=True, color="666666")

    # ── Per-model breakdown ──
    ws.cell(row=11, column=1, value="PER-MODEL BREAKDOWN").font = Font(bold=True, size=13, color="2F5496")

    pm_headers = ["Model", "Mean Baseline (C)", "Mean Future (C)", "Shift (C)", "SCVR (%)"]
    for c, h in enumerate(pm_headers, 1):
        write_label(ws, 12, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    for i, model in enumerate(sorted(baseline.keys())):
        r = 13 + i
        mb = float(np.mean(baseline[model]))
        mf = float(np.mean(future[model]))
        ws.cell(row=r, column=1, value=model)
        write_value(ws, r, 2, mb, fmt=NUM_FMT_2)
        write_value(ws, r, 3, mf, fmt=NUM_FMT_2)
        ws.cell(row=r, column=4, value=f"=C{r}-B{r}").number_format = NUM_FMT_2
        ws.cell(row=r, column=5, value=f"=(C{r}-B{r})/B{r}").number_format = NUM_FMT_PCT

    pm_end = 12 + len(baseline)

    # Ensemble row
    er = pm_end + 1
    write_label(ws, er, 1, "ENSEMBLE (pooled)", font=Font(bold=True), fill=RESULT_FILL)
    write_value(ws, er, 2, mean_b, fmt=NUM_FMT_2, fill=RESULT_FILL)
    write_value(ws, er, 3, mean_f, fmt=NUM_FMT_2, fill=RESULT_FILL)
    ws.cell(row=er, column=4, value=f"=C{er}-B{er}").number_format = NUM_FMT_2
    ws.cell(row=er, column=4).fill = RESULT_FILL
    ws.cell(row=er, column=5, value=f"=(C{er}-B{er})/B{er}").number_format = NUM_FMT_PCT
    ws.cell(row=er, column=5).fill = RESULT_FILL
    ws.cell(row=er, column=5).font = Font(bold=True)

    # ── Bar chart: Area vs Mean comparison ──
    chart = BarChart()
    chart.title = "Baseline vs Future: Area = Mean"
    chart.y_axis.title = "Value (C)"
    chart.style = 10
    chart.width = 20
    chart.height = 14
    chart.type = "col"

    cats = Reference(ws, min_col=1, min_row=5, max_row=6)
    data = Reference(ws, min_col=2, min_row=4, max_row=6, max_col=3)
    chart.add_data(data, titles_from_data=True, from_rows=False)
    chart.set_categories(cats)

    ws.add_chart(chart, "A" + str(er + 3))

    # ── Per-model bar chart ──
    chart2 = BarChart()
    chart2.title = "Per-Model Mean Temperature Shift"
    chart2.y_axis.title = "Mean tasmax (C)"
    chart2.style = 10
    chart2.width = 24
    chart2.height = 14
    chart2.type = "col"

    cats2 = Reference(ws, min_col=1, min_row=13, max_row=pm_end)
    data2 = Reference(ws, min_col=2, min_row=12, max_row=pm_end, max_col=3)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    ws.add_chart(chart2, "G" + str(er + 3))

    auto_width(ws, max_width=22)


def build_experiment_zone(wb, baseline):
    """Sheet 6: Experiment Zone — editable values with live formulas."""
    ws = wb.create_sheet("Experiment Zone")
    ws.sheet_properties.tabColor = "7030A0"

    ws.cell(row=1, column=1, value="EXPERIMENT ZONE: Change values, watch Area = Mean").font = Font(bold=True, size=14, color="7030A0")
    ws.cell(row=2, column=1, value="Instructions: Edit the values in column B. The Area and Mean below will auto-update. They will always match!").font = Font(italic=True, size=11)

    # Use 30 values for experimentation (manageable)
    n_exp = 30
    all_base = np.sort(np.concatenate(list(baseline.values())))[::-1]
    idx = np.linspace(0, len(all_base) - 1, n_exp, dtype=int)
    sample = all_base[idx]
    exc = np.linspace(0, 1, n_exp)

    headers = ["Rank", "Value (C) - EDIT ME!", "Exceedance Prob", "Trap. Width", "Trap. Area"]
    for c, h in enumerate(headers, 1):
        write_label(ws, 4, c, h, font=HEADER_FONT, fill=PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"))

    for i in range(n_exp):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)

        val_cell = ws.cell(row=r, column=2, value=round(float(sample[i]), 2))
        val_cell.number_format = NUM_FMT_2
        val_cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        val_cell.border = THIN_BORDER

        ws.cell(row=r, column=3, value=round(float(exc[i]), 4)).number_format = NUM_FMT_4
        ws.cell(row=r, column=3).border = THIN_BORDER

        if i == 0:
            ws.cell(row=r, column=4, value=0).number_format = NUM_FMT_4
            ws.cell(row=r, column=5, value=0).number_format = NUM_FMT_4
        else:
            ws.cell(row=r, column=4, value=f"=C{r}-C{r-1}").number_format = NUM_FMT_4
            ws.cell(row=r, column=5, value=f"=D{r}*(B{r-1}+B{r})/2").number_format = NUM_FMT_4

        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=5).border = THIN_BORDER

    end_r = 4 + n_exp

    # Results
    sr = end_r + 2
    write_label(ws, sr, 1, "RESULTS", font=Font(bold=True, size=14, color="7030A0"))

    write_label(ws, sr + 1, 1, "Area (SUM of trapezoids):", font=Font(bold=True, size=12), fill=RESULT_FILL)
    area_cell = ws.cell(row=sr + 1, column=2, value=f"=SUM(E5:E{end_r})")
    area_cell.number_format = NUM_FMT_4
    area_cell.fill = RESULT_FILL
    area_cell.font = Font(bold=True, size=14)

    write_label(ws, sr + 2, 1, "Mean (AVERAGE of values):", font=Font(bold=True, size=12), fill=HIGHLIGHT_FILL)
    mean_cell = ws.cell(row=sr + 2, column=2, value=f"=AVERAGE(B5:B{end_r})")
    mean_cell.number_format = NUM_FMT_4
    mean_cell.fill = HIGHLIGHT_FILL
    mean_cell.font = Font(bold=True, size=14)

    write_label(ws, sr + 3, 1, "Do they match?", font=Font(bold=True, size=12))
    ws.cell(row=sr + 3, column=2, value=f'=IF(ABS(B{sr+1}-B{sr+2})<1,"YES!","NO - values must be sorted descending")').font = Font(bold=True, size=14, color="C00000")

    write_label(ws, sr + 5, 1, "TRY THIS:", font=Font(bold=True, size=12, color="7030A0"))
    ws.cell(row=sr + 6, column=1, value="1. Change a few values in column B (keep them sorted high-to-low)").font = Font(size=11)
    ws.cell(row=sr + 7, column=1, value="2. Watch Area and Mean update — they stay equal!").font = Font(size=11)
    ws.cell(row=sr + 8, column=1, value="3. Try making all values the same (e.g. 30) — Area = Mean = 30").font = Font(size=11)
    ws.cell(row=sr + 9, column=1, value="4. Try scrambling the sort order — the match breaks (exceedance requires sorted data)").font = Font(size=11)

    # Chart
    chart = AreaChart()
    chart.title = "Your Exceedance Curve (edit values to change!)"
    chart.x_axis.title = "Exceedance Probability"
    chart.y_axis.title = "Value (C)"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=3, min_row=5, max_row=end_r)
    vals = Reference(ws, min_col=2, min_row=4, max_row=end_r)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "G5")

    auto_width(ws, max_width=35)


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Loading data from NetCDF cache...")
    baseline, future = load_all_data()
    print(f"  Loaded {len(baseline)} models (baseline), {len(future)} models (future)")

    n_base = sum(len(v) for v in baseline.values())
    n_fut = sum(len(v) for v in future.values())
    print(f"  Total: {n_base} baseline days, {n_fut} future days")

    wb = Workbook()

    print("Building sheets...")
    print("  1/6 README")
    build_readme(wb)

    print("  2/6 Raw Daily Data")
    build_raw_data(wb, baseline, future)

    print("  3/6 Exceedance Curve Data")
    build_exceedance_curve(wb, baseline, future)

    print("  4/6 Small Example (20 values)")
    build_small_example(wb, baseline, future)

    print("  5/6 SCVR Calculation")
    build_scvr_calculation(wb, baseline, future)

    print("  6/6 Experiment Zone")
    build_experiment_zone(wb, baseline)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_FILE))
    print(f"\nSaved to: {OUTPUT_FILE}")
    print("Open in Excel / Google Sheets to explore!")


if __name__ == "__main__":
    main()
